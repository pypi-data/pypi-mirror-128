import contextlib
import tempfile

from ..filehandler import FileHandler
from ..util import maybe_convert_usecols, get_version, Scalar
from distutils.version import LooseVersion
from pandas import to_datetime, Timestamp
from typing import List
from numpy import nan
from traceback import format_exc
from logging import getLogger
from copy import deepcopy
from gc import collect

log = getLogger(__name__)


class OpenPYXLReader(FileHandler):
    def __init__(self, file_path, flag_read_only=False, sheet_names=None, header=0, names=None, index_col=None,
                 usecols=None, squeeze=False, dtype=None, true_values=None, false_values=None, skiprows=0, nrows=0,
                 na_values=None, verbose=False, parse_dates=False, date_parser=None, thousands=None, comment=None,
                 skipfooter=0, convert_float=True, mangle_dupe_cols=True, truncate=False, **kwargs):
        super().__init__(file_path=file_path, flag_read_only=flag_read_only, sheet_names=sheet_names, header=header,
                         names=names, index_col=index_col, usecols=usecols, squeeze=squeeze, dtype=dtype,
                         true_values=true_values, false_values=false_values, skiprows=skiprows, nrows=nrows,
                         na_values=na_values, verbose=verbose, parse_dates=parse_dates, date_parser=date_parser,
                         thousands=thousands, comment=comment, skipfooter=skipfooter, convert_float=convert_float,
                         mangle_dupe_cols=mangle_dupe_cols, truncate=truncate, **kwargs)

    def parse(self):
        if self.streams:
            for handler, buffer in self.streams:
                self.__parse(handler=handler, buffer=buffer)
        else:
            return self.__parse()

    def max_rows(self, file_path=None, sheet_names=None):
        sheets = dict()

        if file_path:
            self.file_path = file_path

        if sheet_names:
            self.sheet_names = sheet_names

        if self.file_path:
            try:
                with load_workbook(self.file_path) as xls_file:
                    for tab in xls_file.sheetnames:
                        if not self.sheet_names or tab.lower() in self.sheet_names:
                            xls_ws = xls_file[tab]
                            sheets[tab] = xls_ws.max_row
            except:
                pass
            finally:
                collect()

        return sheets

    def __parse(self, handler=None, buffer=None):
        sheets = dict()

        try:
            with load_workbook(self.file_path, data_only=True) as xls_file:
                if self.flag_read_only and xls_file.read_only:
                    raise ReadOnlyError

                for tab in xls_file.sheetnames:
                    if not self.sheet_names or tab.lower() in self.sheet_names:
                        xls_ws = xls_file[tab]

                        if handler:
                            self.__openpyxl_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
                        else:
                            sheets[tab] = self.__openpyxl_sheet(sheet=xls_ws, handler=handler, buffer=buffer)

                        if self.truncate:
                            try:
                                if xls_ws.max_row > 1:
                                    xls_ws.delete_rows(2, xls_ws.max_row - 1)

                                xls_file.save(filename=file)
                            except Exception as e:
                                log.error(format_exc())
                                log.error('Excel Sheet truncate Failed - Ecode {0}, {1}'.format(type(e).__name__,
                                                                                                str(e)))
                                pass
        except:
            log.error(format_exc())
            pass
        finally:
            collect()

        if sheets:
            return sheets

    def __openpyxl_sheet(self, sheet, handler=None, buffer=None):
        import openpyxl

        version = LooseVersion(get_version(openpyxl))
        is_readonly = hasattr(sheet, "reset_dimensions")

        if version >= "3.0.0" and is_readonly:
            sheet.reset_dimensions()

        data: List[List[Scalar]] = []
        row_number = 0
        last_row_with_data = -1
        header = None

        for row in sheet.rows:
            if 0 <= self.skiprows <= row_number and 0 <= self.header <= row_number and (row_number <= self.nrows
                                                                                        or self.nrows == 0):
                converted_row = [self.__convert_cell(cell, self.convert_float) for cell in row]
                data.append(converted_row)

                if not header:
                    header = converted_row

                if not all(cell == "" for cell in converted_row):
                    last_row_with_data = len(data)

                if handler and buffer < len(data):
                    data = self.__read_only_fix(data, version, is_readonly)
                    df = self.convert_data(data)
                    handler(deepcopy(self.file_path), deepcopy(df), row_number - len(df) + 1,
                            deepcopy(row_number), **self.kwargs)
                    data.clear()
                    data.append(header)
                    last_row_with_data = -1

            row_number += 1

        data = data[: last_row_with_data + 1]
        data = self.__read_only_fix(data, version, is_readonly)
        df = self.convert_data(data)
        data.clear()

        if not df.empty and handler:
            row_number -= 1
            handler(deepcopy(self.file_path), deepcopy(df), row_number - len(df) + 1,
                    deepcopy(row_number), **self.kwargs)
        elif handler is None:
            return df

    @staticmethod
    def __read_only_fix(data, version, is_readonly):
        if version >= "3.0.0" and is_readonly and len(data) > 0:
            max_width = max(len(data_row) for data_row in data)

            if min(len(data_row) for data_row in data) < max_width:
                empty_cell: List[Scalar] = [""]
                data = [
                    data_row + (max_width - len(data_row)) * empty_cell
                    for data_row in data
                ]

        return data

    @staticmethod
    def __convert_cell(cell, convert_float: bool) -> Scalar:
        from openpyxl.cell.cell import TYPE_BOOL, TYPE_ERROR, TYPE_NUMERIC

        if cell.value is None:
            return ""
        elif cell.is_date:
            return cell.value
        elif cell.data_type == TYPE_ERROR:
            return nan
        elif cell.data_type == TYPE_BOOL:
            return bool(cell.value)
        elif cell.data_type == TYPE_NUMERIC:
            if convert_float:
                val = int(cell.value)

                if val == cell.value:
                    return val
            else:
                return float(cell.value)

        return cell.value


class ODFReader(FileHandler):
    __odf_file = None

    def __init__(self, file_path, flag_read_only=False, sheet_names=None, header=0, names=None, index_col=None,
                 usecols=None, squeeze=False, dtype=None, true_values=None, false_values=None, skiprows=0, nrows=0,
                 na_values=None, verbose=False, parse_dates=False, date_parser=None, thousands=None, comment=None,
                 skipfooter=0, convert_float=True, mangle_dupe_cols=True, **kwargs):
        super().__init__(file_path=file_path, flag_read_only=flag_read_only, sheet_names=sheet_names, header=header,
                         names=names, index_col=index_col, usecols=usecols, squeeze=squeeze, dtype=dtype,
                         true_values=true_values, false_values=false_values, skiprows=skiprows, nrows=nrows,
                         na_values=na_values, verbose=verbose, parse_dates=parse_dates, date_parser=date_parser,
                         thousands=thousands, comment=comment, skipfooter=skipfooter, convert_float=convert_float,
                         mangle_dupe_cols=mangle_dupe_cols, **kwargs)

    @property
    def empty_value(self) -> str:
        return ""

    def parse(self):
        if self.streams:
            for handler, buffer in self.streams:
                self.__parse(handler=handler, buffer=buffer)
        else:
            return self.__parse()

    def __parse(self, handler=None, buffer=None):
        from odf.opendocument import load
        from odf.table import Table

        self.__odf_file = load(self.file_path)
        sheets = dict()

        try:
            for xls_ws in self.__odf_file.getElementsByType(Table):
                tab = xls_ws.getAttribute("name")

                if not self.__sheet_names or tab.lower() in self.__sheet_names:
                    if handler:
                        self.__odf_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
                    else:
                        sheets[tab] = self.__odf_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
        finally:
            self.__odf_file.close()
            collect()

        if sheets:
            return sheets

    def __odf_sheet(self, sheet, handler=None, buffer=None):
        from odf.table import CoveredTableCell, TableCell, TableRow

        covered_cell_name = CoveredTableCell().qname
        table_cell_name = TableCell().qname
        cell_names = {covered_cell_name, table_cell_name}

        sheet_rows = sheet.getElementsByType(TableRow)
        header = None
        empty_rows = 0
        max_row_len = 0
        row_number = 0

        table: List[List[Scalar]] = []

        for sheet_row in sheet_rows:
            if 0 < self.skiprows <= row_number <= self.nrows and 0 < self.header <= row_number:
                sheet_cells = [x for x in sheet_row.childNodes if x.qname in cell_names]
                empty_cells = 0
                table_row: List[Scalar] = []

                for j, sheet_cell in enumerate(sheet_cells):
                    if sheet_cell.qname == table_cell_name:
                        value = self.__get_cell_value(sheet_cell)
                    else:
                        value = self.empty_value

                    column_repeat = self.__get_column_repeat(sheet_cell)

                    # Queue up empty values, writing only if content succeeds them
                    if value == self.empty_value:
                        empty_cells += column_repeat
                    else:
                        table_row.extend([self.empty_value] * empty_cells)
                        empty_cells = 0
                        table_row.extend([value] * column_repeat)

                if max_row_len < len(table_row):
                    max_row_len = len(table_row)

                row_repeat = self.__get_row_repeat(sheet_row)

                if self.__is_empty_row(sheet_row):
                    empty_rows += row_repeat
                else:
                    table.extend([[self.empty_value]] * empty_rows)
                    empty_rows = 0

                    for _ in range(row_repeat):
                        table.append(table_row)

                if not header:
                    header = converted_row

                if handler and buffer < len(table):
                    table = self.__fix_table(table)
                    df = self.convert_data(table)
                    handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1),
                            deepcopy(row_number), **self.kwargs)
                    table.clear()
                    table.append(header)

            row_number += 1

        table = self.__fix_table(table)
        df = self.convert_data(table)
        table.clear()

        if not df.empty and handler:
            row_number -= 1
            handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1), deepcopy(row_number),
                    **self.kwargs)
        elif not handler:
            return df

    def __fix_table(self, table):
        for row in table:
            if len(row) < max_row_len:
                row.extend([self.empty_value] * (max_row_len - len(row)))

        return table

    def __get_cell_value(self, cell) -> Scalar:
        from odf.namespaces import OFFICENS

        if str(cell) == "#N/A":
            return nan

        cell_type = cell.attributes.get((OFFICENS, "value-type"))

        if cell_type == "boolean":
            if str(cell) == "TRUE":
                return True

            return False
        if cell_type is None:
            return self.empty_value
        elif cell_type == "float":
            cell_value = float(cell.attributes.get((OFFICENS, "value")))

            if self.convert_float:
                val = int(cell_value)
                if val == cell_value:
                    return val

            return cell_value
        elif cell_type == "percentage":
            cell_value = cell.attributes.get((OFFICENS, "value"))
            return float(cell_value)
        elif cell_type == "string":
            return self.__get_cell_string_value(cell)
        elif cell_type == "currency":
            cell_value = cell.attributes.get((OFFICENS, "value"))
            return float(cell_value)
        elif cell_type == "date":
            cell_value = cell.attributes.get((OFFICENS, "date-value"))
            return to_datetime(cell_value)
        elif cell_type == "time":
            result = to_datetime(str(cell))
            result = cast(Timestamp, result)
            return result.time()
        else:
            self.__odf_file.close()
            raise ValueError(f"Unrecognized type {cell_type}")

    def _get_cell_string_value(self, cell) -> str:
        from odf.element import Element
        from odf.namespaces import TEXTNS
        from odf.text import S

        text_s = S().qname

        value = []

        for fragment in cell.childNodes:
            if isinstance(fragment, Element):
                if fragment.qname == text_s:
                    spaces = int(fragment.attributes.get((TEXTNS, "c"), 1))
                    value.append(" " * spaces)
                else:
                    value.append(self.__get_cell_string_value(fragment))
            else:
                value.append(str(fragment))
        return "".join(value)

    @staticmethod
    def __get_row_repeat(row) -> int:
        from odf.namespaces import TABLENS

        return int(row.attributes.get((TABLENS, "number-rows-repeated"), 1))

    @staticmethod
    def __get_column_repeat(cell) -> int:
        from odf.namespaces import TABLENS

        return int(cell.attributes.get((TABLENS, "number-columns-repeated"), 1))

    @staticmethod
    def __is_empty_row(row) -> bool:
        for column in row.childNodes:
            if len(column.childNodes) > 0:
                return False

        return True


class XLRDReader(FileHandler):
    __xls_file = None

    def __init__(self, file_path, flag_read_only=False, sheet_names=None, header=0, names=None, index_col=None,
                 usecols=None, squeeze=False, dtype=None, true_values=None, false_values=None, skiprows=0, nrows=0,
                 na_values=None, verbose=False, parse_dates=False, date_parser=None, thousands=None, comment=None,
                 skipfooter=0, convert_float=True, mangle_dupe_cols=True, **kwargs):
        super().__init__(file_path=file_path, flag_read_only=flag_read_only, sheet_names=sheet_names, header=header,
                         names=names, index_col=index_col, usecols=usecols, squeeze=squeeze, dtype=dtype,
                         true_values=true_values, false_values=false_values, skiprows=skiprows, nrows=nrows,
                         na_values=na_values, verbose=verbose, parse_dates=parse_dates, date_parser=date_parser,
                         thousands=thousands, comment=comment, skipfooter=skipfooter, convert_float=convert_float,
                         mangle_dupe_cols=mangle_dupe_cols, **kwargs)

    def parse(self):
        if self.streams:
            for handler, buffer in self.streams:
                self.__parse(handler=handler, buffer=buffer)
        else:
            return self.__parse()

    def __parse(self, handler=None, buffer=None):
        from xlrd import open_workbook

        sheets = dict()
        self.__xls_file = open_workbook(self.file_path)

        try:
            for tab in self.__xls_file.sheet_names():
                if not self.sheet_names or tab.lower() in self.sheet_names:
                    xls_ws = self.__xls_file.sheet_by_name(tab)

                    if handler:
                        self.__xlrd_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
                    else:
                        sheets[tab] = self.__xlrd_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
        finally:
            self.__xls_file = None
            collect()

        if sheets:
            return sheets

    def __xlrd_sheet(self, sheet, handler=None, buffer=None):
        data: List[List[Scalar]] = []
        row_number = 0
        header = None

        for i in range(sheet.nrows):
            if 0 < self.skiprows <= row_number <= self.nrows and 0 < self.header <= row_number:
                row = [
                    self.__parse_cell(value, typ)
                    for value, typ in zip(sheet.row_values(i), sheet.row_types(i))
                ]
                data.append(row)

                if not header:
                    header = converted_row

                if handler and len(data) <= buffer:
                    df = self.convert_data(data)
                    handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1),
                            deepcopy(row_number), deepcopy(**self.kwargs))
                    data.clear()
                    data.append(header)

            row_number += 1

        df = self.convert_data(data)
        data.clear()

        if not df.empty and handler:
            row_number -= 1
            handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1),
                    deepcopy(row_number, **self.kwargs))
        elif not handler:
            return df

    def __parse_cell(self, cell_contents, cell_typ):
        from xlrd import (
            XL_CELL_BOOLEAN,
            XL_CELL_DATE,
            XL_CELL_ERROR,
            XL_CELL_NUMBER,
            xldate,
        )

        epoch1904 = self.__xls_file.datemode

        if cell_typ == XL_CELL_DATE:
            try:
                cell_contents = xldate.xldate_as_datetime(cell_contents, epoch1904)
            except OverflowError:
                return cell_contents

            year = (cell_contents.timetuple())[0:3]

            if (not epoch1904 and year == (1899, 12, 31)) or (
                    epoch1904 and year == (1904, 1, 1)
            ):
                cell_contents = time(
                    cell_contents.hour,
                    cell_contents.minute,
                    cell_contents.second,
                    cell_contents.microsecond,
                )
        elif cell_typ == XL_CELL_ERROR:
            cell_contents = np.nan
        elif cell_typ == XL_CELL_BOOLEAN:
            cell_contents = bool(cell_contents)
        elif convert_float and cell_typ == XL_CELL_NUMBER:
            val = int(cell_contents)

            if val == cell_contents:
                cell_contents = val

        return cell_contents


class PYXLSBReader(FileHandler):
    def __init__(self, file_path, flag_read_only=False, sheet_names=None, header=0, names=None, index_col=None,
                 usecols=None, squeeze=False, dtype=None, true_values=None, false_values=None, skiprows=0, nrows=0,
                 na_values=None, verbose=False, parse_dates=False, date_parser=None, thousands=None, comment=None,
                 skipfooter=0, convert_float=True, mangle_dupe_cols=True, **kwargs):
        super().__init__(file_path=file_path, flag_read_only=flag_read_only, sheet_names=sheet_names, header=header,
                         names=names, index_col=index_col, usecols=usecols, squeeze=squeeze, dtype=dtype,
                         true_values=true_values, false_values=false_values, skiprows=skiprows, nrows=nrows,
                         na_values=na_values, verbose=verbose, parse_dates=parse_dates, date_parser=date_parser,
                         thousands=thousands, comment=comment, skipfooter=skipfooter, convert_float=convert_float,
                         mangle_dupe_cols=mangle_dupe_cols, **kwargs)

    def parse(self):
        if self.streams:
            for handler, buffer in self.streams:
                self.__parse(handler=handler, buffer=buffer)
        else:
            return self.__parse()

    def __parse(self, handler=None, buffer=None):
        from pyxlsb import open_workbook

        sheets = dict()
        self.__xls_file = open_workbook(self.file_path)

        try:
            for tab in self.__xls_file.sheets:
                if not self.sheet_names or tab.lower() in self.sheet_names:
                    xls_ws = xls_file.get_sheet(tab)

                    if handler:
                        self.__pyxlsb_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
                    else:
                        sheets[tab] = self.__pyxlsb_sheet(sheet=xls_ws, handler=handler, buffer=buffer)
        finally:
            self.__xls_file.close()
            collect()

        if sheets:
            return sheets

    def __pyxlsb_sheet(self, sheet, handler, buffer):
        data: List[List[Scalar]] = []
        row_number = 0
        header = None

        for r in sheet.rows(sparse=False):
            if 0 < self.skiprows <= row_number <= self.nrows and 0 < self.header <= row_number:
                row = [self.__convert_cell(c) for c in r]
                data.append(row)

                if not header:
                    header = converted_row

                if handler and buffer < len(data):
                    df = self.convert_data(data)
                    handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1),
                            deepcopy(row_number), **self.kwargs)
                    data.clear()
                    data.append(header)

            row_number += 1

        df = self.convert_data(data)
        data.clear()

        if not df.empty and handler:
            row_number -= 1
            handler(deepcopy(self.file_path), deepcopy(df), deepcopy(row_number - len(df) + 1), deepcopy(row_number),
                    **self.kwargs)
        elif not handler:
            return df

    def __convert_cell(self, cell) -> Scalar:
        if cell.v is None:
            return ""
        if isinstance(cell.v, float) and self.convert_float:
            val = int(cell.v)

            if val == cell.v:
                return val
            else:
                return float(cell.v)

        return cell.v


@contextlib.contextmanager
def load_workbook(file_path, *args, **kwargs):
    from openpyxl import load_workbook
    from os.path import join, basename
    from os import remove

    wb = load_workbook(file_path, *args, **kwargs)
    yield wb

    path = join(tempfile.gettempdir(), basename(file_path))
    wb.save(path)
    wb.close()
    remove(path)
    del wb
