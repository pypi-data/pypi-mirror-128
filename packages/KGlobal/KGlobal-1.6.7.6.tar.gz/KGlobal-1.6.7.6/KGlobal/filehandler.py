from pandas import DataFrame, ExcelFile
from os.path import exists, isdir
from xml.etree.ElementTree import parse as xml_parse
from django.utils.crypto import get_random_string
from pandas.core.dtypes.common import is_list_like
from pandas.io.parsers import TextParser
from pandas.errors import EmptyDataError
from typing import Mapping, Any
from .util import DEFAULT_TIMEOUT, DEFAULT_CHECK_INTERVAL, DEFAULT_FAIL_WHEN_LOCKED, LOCK_METHOD


class NotFunctionError(Exception):
    """
        Exception raised for variable not being a callable function
    """
    pass


class Streams(object):
    __engines = ['xmlread', 'csv', 'open', 'xlrd', 'openpyxl', 'odf', 'pyxlsb', 'default']

    __slots__ = "streams"

    def __init__(self):
        self.streams = dict()

    def add_stream(self, engine, handler, buffer):
        assert engine in self.__engines, f"Engine {engine} not recognized"

        if engine not in self.streams.keys():
            self.streams[engine] = list()

        self.streams[engine].append([handler, buffer])

    def clear_streams(self):
        self.streams = dict()

    def keys(self):
        return list(self.streams.keys())

    def items(self):
        return self.streams.items()

    def __getitem__(self, key):
        if isinstance(key, str) and key in self.streams.keys():
            return self.streams[key]
        else:
            return list()

    def __setitem__(self, key, value):
        if isinstance(key, str):
            self.streams[key] = value

    def __delitem__(self, key):
        if isinstance(key, str) and key in self.streams.keys():
            del self.streams[key]

    def __contains__(self, key):
        if isinstance(key, str):
            return key in self.streams.keys()

    def __iter__(self):
        for k in self.streams.keys():
            yield k

    def __len__(self):
        return len(self.streams)

    def __repr__(self):
        return self.__class__.__name__ + repr(str(self.streams))

    def __str__(self):
        return str(self.streams)

    def __eq__(self, other):
        for k in self.__slots__:
            if getattr(self, k) != getattr(other, k):
                return False

        return True


class FileHandler(object):
    __file_path = None
    __file_dir = None
    __sheet_names = list()
    __usecols = None
    __header = 0
    __streams = list()
    __df = DataFrame()
    __fmtparams = dict()

    def __init__(self, file_path=None, file_dir=None, file_name=None, flag_read_only=False, sheet_names=None, header=0,
                 names=None, index_col=None, usecols=None, squeeze=False, dtype=None, true_values=None,
                 false_values=None, skiprows=0, nrows=0, na_values=None, verbose=False, parse_dates=False,
                 date_parser=None, thousands=None, comment=None, skipfooter=0, convert_float=True, dialect=None,
                 mangle_dupe_cols=True, xmlns_rs=None, dict_var=None, df=None, mode='a', timeout=DEFAULT_TIMEOUT,
                 check_interval=DEFAULT_CHECK_INTERVAL, fail_when_locked=DEFAULT_FAIL_WHEN_LOCKED, flags=LOCK_METHOD,
                 truncate=False, fmtparams=None, **kwargs):
        self.file_path = file_path
        self.file_dir = file_dir
        self.file_name = file_name
        self.flag_read_only = flag_read_only
        self.sheet_names = sheet_names
        self.header = header
        self.names = names
        self.index_col = index_col
        self.usecols = usecols
        self.squeeze = squeeze
        self.dtype = dtype
        self.true_values = true_values
        self.false_values = false_values
        self.skiprows = skiprows
        self.nrows = nrows
        self.na_values = na_values
        self.verbose = verbose
        self.parse_dates = parse_dates
        self.date_parser = date_parser
        self.thousands = thousands
        self.comment = comment
        self.skipfooter = skipfooter
        self.convert_float = convert_float
        self.mangle_dupe_cols = mangle_dupe_cols
        self.xmlns_rs = xmlns_rs
        self.dict_var = dict_var
        self.df = df
        self.mode = mode
        self.timeout = timeout,
        self.check_interval = check_interval
        self.fail_when_locked = fail_when_locked
        self.flags = flags
        self.dialect = dialect
        self.truncate = truncate
        self.kwargs = kwargs
        self.fmtparams = fmtparams

    @property
    def fmtparams(self):
        return self.__fmtparams

    @fmtparams.setter
    def fmtparams(self, fmtparams):
        if fmtparams:
            if not isinstance(fmtparams, dict):
                raise ValueError("'fmtparams' %r is not an instance of dict" % fmtparams)

            self.__fmtparams = fmtparams
        else:
            self.__fmtparams = dict()

    @property
    def file_path(self):
        return self.__file_path

    @file_path.setter
    def file_path(self, file_path):
        if file_path:
            if not exists(file_path):
                raise ValueError("'file_path' %r does not exist" % file_path)

            self.__file_path = file_path
        else:
            self.__file_path = None

    @property
    def file_dir(self):
        return self.__file_dir

    @file_dir.setter
    def file_dir(self, file_dir):
        if file_dir:
            if not isdir(file_dir):
                raise ValueError("'file_dir' %r is not a valid directory" % file_dir)
            if not exists(file_dir):
                raise Exception("'xml_file' %r does not exist" % file_dir)

            self.__file_dir = file_dir
        else:
            self.__file_dir = None

    @property
    def sheet_names(self):
        return self.__sheet_names

    @sheet_names.setter
    def sheet_names(self, sheet_names):
        if sheet_names:
            if not isinstance(sheet_names, list):
                raise ValueError("'sheet_names' %r is not a list value" % sheet_name)

            self.__sheet_names = [s.lower() for s in sheet_names]
        else:
            self.__sheet_names = list()

    @property
    def usecols(self):
        return self.__usecols

    @usecols.setter
    def usecols(self, usecols):
        if usecols:
            self.__usecols = maybe_convert_usecols(usecols)
        else:
            self.__usecols = None

    @property
    def header(self):
        return self.__header

    @header.setter
    def header(self, header):
        if header:
            if not isinstance(header, int):
                raise ValueError("'header' %r is not a integer" % header)

            self.__header = header
        else:
            self.__header = 0

    @property
    def streams(self):
        return self.__streams

    @streams.setter
    def streams(self, streams):
        if streams:
            if not isinstance(streams, list):
                raise ValueError("'streams' %r is not a list" % streams)

            self.__streams = streams
        else:
            self.__streams = list()

    @property
    def df(self):
        return self.__df

    @df.setter
    def df(self, df):
        if df:
            if not isinstance(df, DataFrame):
                raise ValueError("'df' %r is not an instance of DataFrame" % df)

            self.__df = df
        else:
            self.__df = DataFrame()

    def convert_data(self, data):
        if data:
            if self.nrows == 0:
                nrows = None
            else:
                nrows = self.nrows

            if is_list_like(self.index_col):
                if self.header is None:
                    offset = 0
                else:
                    offset = 1 + self.header

                if offset < len(data):
                    for col in self.index_col:
                        last = data[offset][col]

                        for row in range(offset + 1, len(data)):
                            if data[row][col] == "" or data[row][col] is None:
                                data[row][col] = last
                            else:
                                last = data[row][col]

            try:
                parser = TextParser(
                    data,
                    names=self.names,
                    header=0,
                    index_col=self.index_col,
                    has_index_names=False,
                    squeeze=self.squeeze,
                    dtype=self.dtype,
                    true_values=self.true_values,
                    false_values=self.false_values,
                    skiprows=None,
                    nrows=nrows,
                    na_values=self.na_values,
                    parse_dates=self.parse_dates,
                    date_parser=self.date_parser,
                    thousands=self.thousands,
                    comment=self.comment,
                    skipfooter=self.skipfooter,
                    usecols=self.usecols,
                    mangle_dupe_cols=self.mangle_dupe_cols,
                    **self.kwargs,
                )

                df = parser.read(nrows=nrows)
            except EmptyDataError:
                df = DataFrame()

            return df
        else:
            return DataFrame()


class FileParser(object):
    """
    File Parser for xml, csv, open, xlrd, excel, and json files
    All formats can be streamed by chunks
    """

    from .reader._xml import XMLReader, XMLWriter
    from .reader._open import OpenReader
    from .reader._excel import OpenPYXLReader, ODFReader, XLRDReader, PYXLSBReader
    from .reader._csv import CSVReader
    from .reader._json import JSONReader

    DEFAULTBUFFER = 1000
    __streams = Streams()
    __reader = None
    __engines: Mapping[str, Any] = {
        "xmlread": XMLReader,
        "xmlwrite": XMLWriter,
        "csv": CSVReader,
        "open": OpenReader,
        "xlrd": XLRDReader,
        "openpyxl": OpenPYXLReader,
        "odf": ODFReader,
        "pyxlsb": PYXLSBReader,
        "json": JSONReader
    }

    @property
    def reader(self):
        return self.__reader

    @reader.setter
    def reader(self, engine_or_reader):
        if engine_or_reader:
            if isinstance(engine_or_reader, str):
                assert engine_or_reader in self.__engines, f"Engine {engine_or_reader} not recognized"
                engine_or_reader = self.__engines[engine_or_reader]

            self.__reader = engine_or_reader
        else:
            self.__reader = None

    def stream_data(self, buffer=DEFAULTBUFFER, engine='default'):
        """
        Stream data into chunks

        :param buffer: Number of lines per chunk
        :param engine: [Optional] xmlread, csv, open, xlrd, openpyxl, odf, pyxlsb, or default
        :var data: Either dataframe or list of data
        :var row_start: Row number for first row of data
        :var row_end: Row number for last row of data
        :var **kwargs: Kwarg values provided by parse commands
        """

        def registerhandler(handler):
            self.__streams.add_stream(engine, handler, buffer)
            return handler

        return registerhandler

    def clear_streams(self):
        self.__streams.clear_streams()

    def parse_excel(self, file_path, flag_read_only=False, sheet_names=None, header=0, names=None, index_col=None,
                    usecols=None, squeeze=False, converters=None, dtype=None, engine=None, true_values=None,
                    false_values=None, skiprows=0, nrows=0, na_values=None, keep_default_na=True, na_filter=True,
                    verbose=False, parse_dates=False, date_parser=None, thousands=None, comment=None,
                    convert_float=True, mangle_dupe_cols=True, truncate=False, **kwargs):

        """
        Parses excel files into a dataframe or stream

        :param file_path: File path to the file being read
        :param flag_read_only: [Optional] (True/False) Creates an error if file is opened in read-only mode
        :param sheet_names: List of sheet names to read
        :param header: Row number that is the header (All preceding rows will be truncated)
        :param names: List of column names to use
        :param index_col: Column(s) to index
        :param usecols: Columns to parse (i.e "A:E")
        :param squeeze: Fit data into a Series if one column
        :param converters: Dict of functions for converting values in certain columns
        :param dtype: Set data types for specific columns
        :param engine: Excel engine to use
        :param true_values: Values to consider as True
        :param false_values: Values to consider as False
        :param skiprows: Rows to skip
        :param nrows: Number of rows to grab
        :param na_values: Additional strings to recognize as NA/NaN
        :param keep_default_na: Whether or not to include the default NaN values when parsing the data
        :param na_filter: Detect missing value markers
        :param verbose: Indicate number of NA values placed in non-numeric columns
        :param parse_dates: The behavior is as follows:
            bool. If True -> try parsing the index.
            list of int or names. e.g. If [1, 2, 3] -> try parsing columns 1, 2, 3 each as a separate date column.
            list of lists. e.g. If [[1, 3]] -> combine columns 1 and 3 and parse as a single date column.
            dict, e.g. {‘foo’ : [1, 3]} -> parse columns 1, 3 as date and call result ‘foo’
        :param date_parser: Function to use for converting a sequence of string columns to an array of datetime instances
        :param thousands: Thousands separator for parsing string columns to numeric
        :param comment: Comments out remainder of line
        :param convert_float: Convert integral floats to int (i.e., 1.0 –> 1)
        :param mangle_dupe_cols: Duplicate columns will be specified as ‘X’, ‘X.1’, …’X.N’, rather than ‘X’…’X’
        :param truncate: Truncates excel file
        :param kwargs: kwarg values
        :return: Dataframe
        """

        obj = ExcelFile(file_path, engine)
        self.reader = self.__engines[obj.engine](file_path=file_path, flag_read_only=flag_read_only,
                                                 sheet_names=sheet_names, header=header, names=names,
                                                 index_col=index_col, usecols=usecols, squeeze=squeeze, dtype=dtype,
                                                 true_values=true_values, false_values=false_values, skiprows=skiprows,
                                                 nrows=nrows, na_values=na_values, verbose=verbose,
                                                 parse_dates=parse_dates, date_parser=date_parser, thousands=thousands,
                                                 comment=comment, skipfooter=0, convert_float=convert_float,
                                                 mangle_dupe_cols=mangle_dupe_cols, converters=converters,
                                                 keep_default_na=keep_default_na, na_filter=na_filter,
                                                 truncate=truncate, **kwargs)
        self.reader.streams = self.__streams[obj.engine] + self.__streams['default']
        return self.reader.parse()

    def parse_xml(self, file_path, xmlns_rs, dict_var=None, **kwargs):
        """
        Parses XML file into dataframe or stream

        :param file_path: File path to the file being read
        :param xmlns_rs: xmlns_rs header needed to read xml file (At top of xml file)
        :param dict_var: [Optional] (dict) Append data into dict variable
        :param kwargs: kwarg values
        :return: dataframe
        """

        self.reader = self.__engines['xmlread'](file_path=file_path, xmlns_rs=xmlns_rs, dict_var=dict_var, **kwargs)
        self.reader.streams = self.__streams['xmlread'] + self.__streams['default']
        return self.reader.parse()

    def write_xml(self, file_dir, file_name, df):
        """
        Writes XML data into file

        :param file_dir: File directory where to XML file at
        :param file_name: Filename of XML file
        :param df: Dataframe to write into XML file
        """

        self.reader = self.__engines['xmlwriter'](file_dir=file_dir, file_name=file_name, df=df)
        self.reader.write()

    def parse_open(self, file_path, mode='a', timeout=DEFAULT_TIMEOUT, check_interval=DEFAULT_CHECK_INTERVAL,
                   fail_when_locked=DEFAULT_FAIL_WHEN_LOCKED, flags=LOCK_METHOD, **kwargs):
        """
        Opens a Txt or etc... file and returns a file open object or streams into buffered data

        :param file_path: File path to the file being read
        :param mode: File mode (ie r, rb, w, a, etc...)
        :param timeout: Timeout length
        :param check_interval: Check interval
        :param fail_when_locked: Type of failure when locked
        :param flags: Flags
        :param kwargs: kwarg values
        :return: open object for reading
        """

        self.reader = self.__engines['open'](file_path=file_path, mode=mode, timeout=timeout,
                                             check_interval=check_interval, fail_when_locked=fail_when_locked,
                                             flags=flags, **kwargs)
        self.reader.streams = self.__streams['open'] + self.__streams['default']
        return self.reader.parse()

    def parse_csv(self, file_path, dialect=None, truncate=False, fmtparams=None, **kwargs):
        """
        Parses CSV data into data

        :param file_path: File path to the file being read
        :param dialect: dialect for csv reading
        :param truncate: Truncates excel file
        :param kwargs: Python format parameters
        :return: data
        """

        self.reader = self.__engines['csv'](file_path=file_path, dialect=dialect, truncate=truncate,
                                            fmtparams=fmtparams, **kwargs)
        self.reader.streams = self.__streams['csv'] + self.__streams['default']
        return self.reader.parse()

    def parse_json(self, file_path, **kwargs):
        """
        Reads json into python objects or stream it in chunks

        :param file_path: File path to the file being read
        :param kwargs: kwarg values
        :return: python objects
        """

        self.reader = self.__engines['json'](file_path=file_path, **kwargs)
        self.reader.streams = self.__streams['json'] + self.__streams['default']
        return self.reader.parse()

    def __del__(self):
        self.clear_streams()
        self.__streams = None
        self.__reader = None


def is_an_xml(xml_file):
    """
    Checks to see if XML file is a legit csv file

    :param xml_file: XML File path
    :return: True/False
    """

    if not exists(xml_file):
        raise Exception("'xml_file' %r does not exist" % xml_file)

    try:
        xml_parse(xml_file)
        return True
    except:
        pass

    return False


def unique_id():
    """
    Creates a unique identifier

    :return: Unique identifier
    """

    return get_random_string(length=16,
                             allowed_chars=u'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789')


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
